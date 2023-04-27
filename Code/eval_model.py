import argparse
import numpy as np
np.random.seed(3)
from sklearn.metrics import classification_report
import matplotlib.pyplot as plt
import seaborn as sns
sns.set(rc={"figure.figsize": (6, 4), "figure.dpi": 150})
sns.set_style("white")
import csv
import os
import shutil
import sys
import tarfile
import zipfile
import datetime
import pandas as pd

from sklearn.metrics import roc_curve
from sklearn.metrics import auc
from sklearn.metrics import ConfusionMatrixDisplay
from sklearn.metrics import confusion_matrix

from preprocessing import calculate_weights,load_data, oversample_data

from openpyxl import load_workbook, Workbook

# helper function to get data_path
def set_test_data_path(test_dataset, flatten=False):
    if flatten:
        if test_dataset == "1":
            test_data_path = "/content/drive/MyDrive/Paper3Logging/Datasets/dataset_1_flattened.zip"
        elif test_dataset == "2":
            test_data_path = "/content/drive/MyDrive/Paper3Logging/Datasets/dataset_2_flattened.zip"
        elif test_dataset == "3":
            test_data_path = "/content/drive/MyDrive/Paper3Logging/Datasets/dataset_3_flattened.zip"
        elif test_dataset == "4":
            test_data_path = "/content/drive/MyDrive/Paper3Logging/Datasets/dataset_4.zip"
        else:
            raise ValueError("Invalid test dataset value")
        return test_data_path
    else:
        if test_dataset == "1":
            test_data_path = "/content/drive/MyDrive/Paper3Logging/Datasets/dataset_1.zip"
        elif test_dataset == "2":
            test_data_path = "/content/drive/MyDrive/Paper3Logging/Datasets/dataset_2.zip"
        elif test_dataset == "3":
            test_data_path = "/content/drive/MyDrive/Paper3Logging/Datasets/dataset_3.zip"
        elif test_dataset == "4":
            test_data_path = "/content/drive/MyDrive/Paper3Logging/Datasets/dataset_4.zip"
        else:
            raise ValueError("Invalid test dataset value")
        return test_data_path

def fetch_data(test_dataset,test_data_path):
    # name for local dir
    filename = "dataset_{}".format(test_dataset)
    # Create directory for extracted files
    extract_dir = os.path.join(os.getcwd(), filename)
    os.makedirs(extract_dir, exist_ok=True)
    # uncompress files
    if test_data_path.endswith('.zip'):
        with zipfile.ZipFile(test_data_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
    elif test_data_path.endswith('.tar'):
        with tarfile.open(test_data_path, 'r') as tar_ref:
            tar_ref.extractall(extract_dir)
    else:
        raise ValueError('Test data file type not recognized. Please use .zip or .tar files.')
    print('Data fetched and extracted successfully.')
from decimal import *

def calcSensitivityThresh(targetSensitivity, covidProb,normalProb,sens98=0):
  def find_sens_spec( covid_prob, noncovid_prob, thresh):
    sensitivity= (covid_prob >= thresh).sum()   / (len(covid_prob)+1e-10)
    specificity= (noncovid_prob < thresh).sum() / (len(noncovid_prob)+1e-10)
    return sensitivity, specificity
  getcontext().prec = 3
  thresh = Decimal(sens98)
  sens = 1

  while sens>targetSensitivity:
    thresh = thresh + Decimal(0.1)
    sens,spec = find_sens_spec( covidProb, normalProb,thresh)
    # print("sens: {}   thresh: {}".format(sens,thresh))
  while sens<targetSensitivity:
    thresh = thresh - Decimal(0.01)
    sens,spec = find_sens_spec( covidProb, normalProb,thresh)
    # print("sens: {}   thresh: {}".format(sens,thresh))
  while np.round(sens,3)>=targetSensitivity:
    thresh = thresh + Decimal(0.001)
    sens,spec = find_sens_spec( covidProb, normalProb,thresh)
    # print("sens: {}   thresh: {}".format(sens,thresh))
  if np.round(sens,3)!= targetSensitivity:
    thresh = thresh - Decimal(0.001)
    sens,spec = find_sens_spec( covidProb, normalProb,thresh)
  
  print("Threshold = %.3f" %(thresh))

def getSensitivityThresh(targetSensitivity, covidProb,normalProb,sens98=0):
  def find_sens_spec( covid_prob, noncovid_prob, thresh):
    sensitivity= (covid_prob >= thresh).sum()   / (len(covid_prob)+1e-10)
    specificity= (noncovid_prob < thresh).sum() / (len(noncovid_prob)+1e-10)
    return sensitivity, specificity
  thresh = sens98
  sens = 1
  while sens>targetSensitivity:
    thresh = thresh + 0.1
    sens,spec = find_sens_spec( covidProb, normalProb,thresh)
  while sens<targetSensitivity:
    thresh = thresh - 0.01
    sens,spec = find_sens_spec( covidProb, normalProb,thresh)
  while np.round(sens,3)>=targetSensitivity:
    thresh = thresh + 0.001
    sens,spec = find_sens_spec( covidProb, normalProb,thresh)
  if np.round(sens,3)!= targetSensitivity:
    thresh = thresh - 0.001
    sens,spec = find_sens_spec( covidProb, normalProb,thresh)
  print("Threshold= %.3f" %(thresh))
  return thresh
# ========================================================================================================
# Set up base parameters
parser = argparse.ArgumentParser(description='Evaluation tool')

parser.add_argument('-mod', '--model_path', help='Model Name: ', required=True)
parser.add_argument('-res', '--img_res', help='Image resolution', default=224, type=int)

parser.add_argument('-b', '--batch_size', help='Batch size', default=32, type=int)

parser.add_argument('-ds1', '--dataset1', help='Use Dataset 1 - Minaee Covid', action='store_true', default=False)
parser.add_argument('-ds2', '--dataset2', help='Use Dataset 2 - Cohen + Brax', action='store_true', default=False)
parser.add_argument('-ds3', '--dataset3', help='Use Dataset 3 - SIIM', action='store_true', default=False)
parser.add_argument('-ds4', '--dataset4', help='Use Dataset 4 - COVIDGR', action='store_true', default=False)

parser.add_argument('--no_export', help='Disable result exporting', action='store_true', default=False)

args = parser.parse_args()

# ========================================================================================================
# parse data param

eval_datasets = []
if args.dataset1:
    eval_datasets.append("1")
if args.dataset2:
    eval_datasets.append("2")
if args.dataset3:
    eval_datasets.append("3")
if args.dataset4:
    eval_datasets.append("4")
if eval_datasets==[]:
    raise ValueError("No eval dataset selected")

# set env params from argparse
model_path = args.model_path
img_res = args.img_res
batch_size = args.batch_size
augmented = False
no_export = args.no_export

# ========================================================================================================
# TEMP
temp_flatten = True

architecture = model_path.split("/")[-1].split("_")[0]
model_name = model_path.split("/")[-1]
model_name_dir = model_name.split(".h5")[0][:-2]

trained_on_dataset = model_name_dir.split("_")[2]


# ========================================================================================================
    # loading models 

if architecture == "darkcovidnet":
    pass
elif architecture == "minaee-resnet":
    pass
else:
    # Use TensorFlow code
    import tensorflow as tf
    ## GLOBAL SEED ##                                                   
    tf.random.set_seed(3)
    
    from tensorflow import keras
    from keras import layers
    from keras import models

    print("Num GPUs Available: ", len(tf.config.experimental.list_physical_devices('GPU')))

    # load model
    model =keras.models.load_model(model_path)

    # fetch  data into environment
    for ds in eval_datasets:
        # extract data locally
        fetch_data(ds,set_test_data_path(ds,temp_flatten))
        # local verison path
        data_path = "dataset_{}".format(ds)
        print("Evalating on {}".format(data_path))

        # check if model needs preprocessing (from name)
        hist_eq = True if "_HE_" in model_name else False

        # load in test data
        test_x,test_y, test_paths = load_data(os.path.join(data_path,"test"),img_res,hist_eq)
        
        # eval dataset
        preds = model.predict(test_x)

        # get model number for preds number
        model_number =  model_name.split(".h5")[0][-1]

        # Create folder for preds
        preds_folder = os.path.join(model_path.split(model_name)[0],"Predictions")
        if not no_export:
            os.makedirs(preds_folder, exist_ok=True)
        
        # dataset preds for this model on this dataset
        data_pred_file = os.path.join(preds_folder,data_path+".csv")

        if os.path.isfile(data_pred_file):
          data_df = pd.read_csv(data_pred_file)
          data_df['Pred{}'.format(model_number)] = preds.ravel()
        else:
          # create new csv for storing model meta data
          data_df = pd.DataFrame(columns=['Labels', 'Pred1', 'Pred2','Pred3'])
          data_df['Labels'] = test_y
          data_df['Pred{}'.format(model_number)] = preds.ravel()

        if not no_export:
            data_df.to_csv(data_pred_file,index=False)

        # calculate metrics
        fpr,tpr,thresholds = roc_curve(data_df['Labels'],data_df['Pred{}'.format(model_number)],drop_intermediate=False)
        model_auc = auc(fpr,tpr)

        # if internal eval, calc 98% metrics
        if trained_on_dataset == ds:
            # 98
            print("Internal Evaluation")
            covid_prob= data_df[data_df['Labels']==1]['Pred{}'.format(model_number)].array
            non_prob= data_df[data_df['Labels']==0]['Pred{}'.format(model_number)].array
            thresh = getSensitivityThresh(0.98,covid_prob, non_prob)
        else:
            # calc j stat for optimal thresh https://machinelearningmastery.com/threshold-moving-for-imbalanced-classification/
            print("External Evaluation")
            thresh = thresholds[np.argmax(tpr - fpr)]
        
        cm = confusion_matrix(data_df['Labels'],data_df['Pred{}'.format(model_number)].apply(lambda x: 0 if x<=thresh else 1))

        TP = cm[1][1]
        TN = cm[0][0]
        FP = cm[0][1]
        FN = cm[1][0]

        precision = TP/(TP+FP)
        prec_samp = TP+FP
        sensitivity = TP/(TP+FN)
        specificity = TN/(TN+FP)

        print("AUC: {}".format(model_auc))
        print("Sens: {:.1f}% \nSpec: {:.1f}% \nPrec: {:.1f}%".format(\
            sensitivity*100,specificity*100,precision*100))
        if not no_export:
            # export results
            wb_path = "drive/MyDrive/Paper3Logging/Models/Results.xlsx"
            if os.path.isfile(wb_path):
                # load workbook
                wb = load_workbook(wb_path)
            else:
                wb = Workbook()
                # create workbook
                sheet1 = wb.active
                sheet1.title = "Dataset1"
                sheet2 = wb.create_sheet(title="Dataset2")
                sheet3 = wb.create_sheet(title="Dataset3")
                sheet4 = wb.create_sheet(title="Dataset4")
                headers = ["Model_Name","Eval_Type","AUC","Sensitivity",\
                "Specificity","Precision"]
                sheet1.append(headers)
                sheet2.append(headers)
                sheet3.append(headers)
                sheet4.append(headers)
            # get sheet to update
            sheet = wb['Dataset{}'.format(ds)]
            eval_type = "INTERNAL" if trained_on_dataset==ds else "EXTERNAL"
            sheet.append([model_name, eval_type, model_auc, sensitivity, specificity, precision])
            wb.save(filename=wb_path)

